# app/services/tracker.py
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Dict, cast

from openpyxl import load_workbook
from openpyxl.workbook import Workbook as XLWorkbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook as XLNewWorkbook  # for leads file

from typing import cast
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook as XLNewWorkbook
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime

# ----- Applications tracker -----

HEADERS = [
    "company", "role", "date_applied", "job_url", "source", "ats_type",
    "confirmation_number", "status", "resume_version", "notes",
]
SHEET_NAME = "Applications"


def _ensure_workbook(path: Path) -> None:
    """Create workbook + sheet + headers if missing; normalize header row."""
    if not path.exists():
        wb = XLWorkbook()
        ws = cast(Worksheet, wb.active)
        ws.title = SHEET_NAME
        ws.append(HEADERS)
        wb.save(str(path))
        return

    wb = load_workbook(str(path))
    if SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_NAME)
        ws.append(HEADERS)
        wb.save(str(path))
    else:
        ws = wb[SHEET_NAME]
        if ws.max_row == 1 and [c.value for c in ws[1]] != HEADERS:
            ws.delete_rows(1, ws.max_row)
            ws.append(HEADERS)
            wb.save(str(path))


def log_to_excel(xlsx_path: str | Path, row: Dict[str, Any]) -> str:
    """Append an application entry and return the XLSX absolute path."""
    p = Path(xlsx_path).resolve()
    p.parent.mkdir(parents=True, exist_ok=True)
    _ensure_workbook(p)

    wb = load_workbook(str(p))
    ws: Worksheet = wb[SHEET_NAME]
    ws.append(
        [
            row.get("company", ""),
            row.get("role", ""),
            row.get("date_applied", datetime.utcnow().isoformat(timespec="seconds")),
            row.get("job_url", ""),
            row.get("source", ""),
            row.get("ats_type", ""),
            row.get("confirmation_number", ""),
            row.get("status", ""),
            row.get("resume_version", ""),
            row.get("notes", ""),
        ]
    )
    wb.save(str(p))
    return str(p)

# ----- Leads list (picked jobs) -----



def log_leads_to_excel(path: str | Path, rows: list[dict]) -> str:
    """
    Ensure a leads workbook exists and append rows.
    Expected keys per row: company, title, job_url, ats_type, imported_at (optional)
    """
    p = Path(path).resolve()
    p.parent.mkdir(parents=True, exist_ok=True)

    # Create the file with a "Leads" sheet + header if it doesn't exist yet
    if not p.exists():
        wb = XLNewWorkbook()
        ws: Worksheet = cast(Worksheet, wb.active)   # <-- important cast
        ws.title = "Leads"
        ws.append(["company", "title", "job_url", "ats_type", "imported_at"])
        wb.save(str(p))

    # Open and get the "Leads" sheet (create if missing)
    wb = load_workbook(str(p))
    if "Leads" not in wb.sheetnames:
        ws_created = wb.create_sheet("Leads")
        ws: Worksheet = cast(Worksheet, ws_created)
        ws.append(["company", "title", "job_url", "ats_type", "imported_at"])
    else:
        ws = cast(Worksheet, wb["Leads"])

    # Append rows
    for r in rows:
        ws.append([
            r.get("company", ""),
            r.get("title", ""),
            r.get("job_url", ""),
            r.get("ats_type", ""),
            r.get("imported_at", datetime.utcnow().isoformat(timespec="seconds")),
        ])

    wb.save(str(p))
    return str(p)
